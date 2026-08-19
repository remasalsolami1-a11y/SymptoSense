import os
import io
import json
import logging
import sqlite3
import hashlib
from datetime import datetime, timedelta, timezone
from collections import Counter

# PostgreSQL (production, Railway) if DATABASE_URL is set, otherwise SQLite (local dev).
DB_PATH = os.environ.get("DB_PATH", "symptosense.db")
DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
USE_POSTGRES = bool(DATABASE_URL)

PH = "%s" if USE_POSTGRES else "?"

_logger = logging.getLogger("SymptoSense")


def _init_backend():
    """Validates the PostgreSQL connection once. On failure, falls back to SQLite
    so the bot keeps running instead of crashing, and logs the real error."""
    global USE_POSTGRES, PH
    if not USE_POSTGRES:
        return
    try:
        import psycopg2
        conn = psycopg2.connect(DATABASE_URL)
        conn.close()
    except Exception as e:
        _logger.error(
            f"PostgreSQL connection failed ({e!r}) — falling back to SQLite. "
            f"Fix DATABASE_URL to enable persistent storage."
        )
        USE_POSTGRES = False
        PH = "?"


def _conn():
    if USE_POSTGRES:
        import psycopg2
        return psycopg2.connect(DATABASE_URL)
    os.makedirs(os.path.dirname(os.path.abspath(DB_PATH)), exist_ok=True)
    return sqlite3.connect(DB_PATH)


def init_db():
    _init_backend()
    conn = _conn()
    try:
        c = conn.cursor()
        if USE_POSTGRES:
            c.execute("""
                CREATE TABLE IF NOT EXISTS records (
                    id SERIAL PRIMARY KEY,
                    user_hash TEXT NOT NULL,
                    timestamp TEXT NOT NULL,
                    lang TEXT,
                    age INTEGER,
                    gender TEXT,
                    symptoms TEXT,
                    conditions TEXT,
                    medications TEXT,
                    duration TEXT,
                    severity INTEGER,
                    urgency TEXT
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_user_hash ON records(user_hash)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_timestamp ON records(timestamp)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS visits (
                    id SERIAL PRIMARY KEY,
                    user_hash TEXT NOT NULL,
                    timestamp TEXT NOT NULL
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_visit_user ON visits(user_hash)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_visit_timestamp ON visits(timestamp)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS subscribers (
                    user_id BIGINT PRIMARY KEY,
                    lang TEXT DEFAULT 'ar',
                    subscribed INTEGER DEFAULT 1,
                    added_at TEXT
                )
            """)
            c.execute("""
                CREATE TABLE IF NOT EXISTS followups (
                    id SERIAL PRIMARY KEY,
                    user_hash TEXT NOT NULL,
                    record_id INTEGER,
                    timestamp TEXT NOT NULL,
                    outcome TEXT
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_fu_record ON followups(record_id)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS daily_checkins (
                    id SERIAL PRIMARY KEY,
                    user_hash TEXT NOT NULL,
                    timestamp TEXT NOT NULL,
                    severity INTEGER
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_ci_user ON daily_checkins(user_hash)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS med_reminders (
                    id SERIAL PRIMARY KEY,
                    user_id BIGINT,
                    med_name TEXT,
                    time_utc TEXT,
                    lang TEXT DEFAULT 'ar',
                    active INTEGER DEFAULT 1,
                    created TEXT
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_med_user ON med_reminders(user_id)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS feedback (
                    id SERIAL PRIMARY KEY,
                    user_hash TEXT NOT NULL,
                    record_id INTEGER,
                    rating TEXT,
                    comment TEXT,
                    timestamp TEXT NOT NULL
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_fb_record ON feedback(record_id)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS assistant_feedback (
                    id SERIAL PRIMARY KEY,
                    user_hash TEXT NOT NULL,
                    message TEXT,
                    rating INTEGER,
                    reason TEXT,
                    timestamp TEXT NOT NULL
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_asst_fb_time ON assistant_feedback(timestamp)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS user_data (
                    user_id BIGINT PRIMARY KEY,
                    data TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                )
            """)
            c.execute("""
                CREATE TABLE IF NOT EXISTS conversations (
                    name TEXT NOT NULL,
                    key TEXT NOT NULL,
                    state TEXT,
                    PRIMARY KEY (name, key)
                )
            """)
            c.execute("""
                CREATE TABLE IF NOT EXISTS blood_tests (
                    id SERIAL PRIMARY KEY,
                    user_hash TEXT NOT NULL,
                    timestamp TEXT NOT NULL,
                    data TEXT NOT NULL
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_blood_user ON blood_tests(user_hash)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS profiles (
                    user_hash TEXT PRIMARY KEY,
                    lang TEXT DEFAULT 'ar',
                    age TEXT,
                    gender TEXT,
                    conditions TEXT,
                    medications TEXT,
                    allergies TEXT,
                    updated_at TEXT NOT NULL
                )
            """)
            c.execute("""
                CREATE TABLE IF NOT EXISTS results (
                    user_hash TEXT NOT NULL,
                    record_id INTEGER NOT NULL,
                    data TEXT NOT NULL,
                    PRIMARY KEY (user_hash, record_id)
                )
            """)
            c.execute("""
                CREATE TABLE IF NOT EXISTS family_members (
                    id SERIAL PRIMARY KEY,
                    user_hash TEXT NOT NULL,
                    relation TEXT,
                    name TEXT NOT NULL,
                    age TEXT,
                    gender TEXT,
                    conditions TEXT,
                    medications TEXT,
                    allergies TEXT,
                    notes TEXT,
                    created TEXT
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_fam_user ON family_members(user_hash)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS med_plans (
                    id SERIAL PRIMARY KEY,
                    user_hash TEXT NOT NULL,
                    member_id INTEGER DEFAULT 0,
                    med_name TEXT NOT NULL,
                    dose TEXT,
                    times TEXT NOT NULL,
                    frequency TEXT DEFAULT 'daily',
                    start_date TEXT,
                    days INTEGER,
                    active INTEGER DEFAULT 1,
                    created TEXT
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_mp_user ON med_plans(user_hash)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS med_logs (
                    id SERIAL PRIMARY KEY,
                    user_hash TEXT NOT NULL,
                    member_id INTEGER DEFAULT 0,
                    plan_id INTEGER NOT NULL,
                    log_date TEXT NOT NULL,
                    log_time TEXT NOT NULL,
                    status TEXT NOT NULL,
                    created TEXT
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_ml_plan ON med_logs(plan_id, log_date)")
        else:
            c.execute("""
                CREATE TABLE IF NOT EXISTS records (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_hash TEXT NOT NULL,
                    timestamp TEXT NOT NULL,
                    lang TEXT,
                    age INTEGER,
                    gender TEXT,
                    symptoms TEXT,
                    conditions TEXT,
                    medications TEXT,
                    duration TEXT,
                    severity INTEGER,
                    urgency TEXT
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_user_hash ON records(user_hash)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_timestamp ON records(timestamp)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS visits (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_hash TEXT NOT NULL,
                    timestamp TEXT NOT NULL
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_visit_user ON visits(user_hash)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_visit_timestamp ON visits(timestamp)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS subscribers (
                    user_id INTEGER PRIMARY KEY,
                    lang TEXT DEFAULT 'ar',
                    subscribed INTEGER DEFAULT 1,
                    added_at TEXT
                )
            """)
            c.execute("""
                CREATE TABLE IF NOT EXISTS followups (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_hash TEXT NOT NULL,
                    record_id INTEGER,
                    timestamp TEXT NOT NULL,
                    outcome TEXT
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_fu_record ON followups(record_id)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS daily_checkins (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_hash TEXT NOT NULL,
                    timestamp TEXT NOT NULL,
                    severity INTEGER
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_ci_user ON daily_checkins(user_hash)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS med_reminders (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER,
                    med_name TEXT,
                    time_utc TEXT,
                    lang TEXT DEFAULT 'ar',
                    active INTEGER DEFAULT 1,
                    created TEXT
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_med_user ON med_reminders(user_id)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS feedback (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_hash TEXT NOT NULL,
                    record_id INTEGER,
                    rating TEXT,
                    comment TEXT,
                    timestamp TEXT NOT NULL
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_fb_record ON feedback(record_id)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS assistant_feedback (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_hash TEXT NOT NULL,
                    message TEXT,
                    rating INTEGER,
                    reason TEXT,
                    timestamp TEXT NOT NULL
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_asst_fb_time ON assistant_feedback(timestamp)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS user_data (
                    user_id INTEGER PRIMARY KEY,
                    data TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                )
            """)
            c.execute("""
                CREATE TABLE IF NOT EXISTS conversations (
                    name TEXT NOT NULL,
                    key TEXT NOT NULL,
                    state TEXT,
                    PRIMARY KEY (name, key)
                )
            """)
            c.execute("""
                CREATE TABLE IF NOT EXISTS blood_tests (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_hash TEXT NOT NULL,
                    timestamp TEXT NOT NULL,
                    data TEXT NOT NULL
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_blood_user ON blood_tests(user_hash)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS profiles (
                    user_hash TEXT PRIMARY KEY,
                    lang TEXT DEFAULT 'ar',
                    age TEXT,
                    gender TEXT,
                    conditions TEXT,
                    medications TEXT,
                    allergies TEXT,
                    updated_at TEXT NOT NULL
                )
            """)
            c.execute("""
                CREATE TABLE IF NOT EXISTS results (
                    user_hash TEXT NOT NULL,
                    record_id INTEGER NOT NULL,
                    data TEXT NOT NULL,
                    PRIMARY KEY (user_hash, record_id)
                )
            """)
            c.execute("""
                CREATE TABLE IF NOT EXISTS family_members (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_hash TEXT NOT NULL,
                    relation TEXT,
                    name TEXT NOT NULL,
                    age TEXT,
                    gender TEXT,
                    conditions TEXT,
                    medications TEXT,
                    allergies TEXT,
                    notes TEXT,
                    created TEXT
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_fam_user ON family_members(user_hash)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS med_plans (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_hash TEXT NOT NULL,
                    member_id INTEGER DEFAULT 0,
                    med_name TEXT NOT NULL,
                    dose TEXT,
                    times TEXT NOT NULL,
                    frequency TEXT DEFAULT 'daily',
                    start_date TEXT,
                    days INTEGER,
                    active INTEGER DEFAULT 1,
                    created TEXT
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_mp_user ON med_plans(user_hash)")
            c.execute("""
                CREATE TABLE IF NOT EXISTS med_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_hash TEXT NOT NULL,
                    member_id INTEGER DEFAULT 0,
                    plan_id INTEGER NOT NULL,
                    log_date TEXT NOT NULL,
                    log_time TEXT NOT NULL,
                    status TEXT NOT NULL,
                    created TEXT
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_ml_plan ON med_logs(plan_id, log_date)")
        # ---- Web Push notification tables ----
        _create_push_tables(c)
        # ---- Smart Account System tables ----
        _create_ss_tables(c)
        conn.commit()
        _migrate_records(conn, c)
        _migrate_feedback(conn, c)
        _migrate_members(conn, c)
        _migrate_ss_columns(conn, c)
        conn.commit()
    finally:
        conn.close()


def _create_push_tables(c):
    """Create Web Push subscription and delivery de-duplication tables."""
    if USE_POSTGRES:
        c.execute("""
            CREATE TABLE IF NOT EXISTS push_subscriptions (
                id SERIAL PRIMARY KEY,
                user_hash TEXT NOT NULL,
                endpoint TEXT UNIQUE NOT NULL,
                p256dh TEXT NOT NULL,
                auth TEXT NOT NULL,
                timezone TEXT DEFAULT 'Asia/Riyadh',
                lang TEXT DEFAULT 'ar',
                active INTEGER DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        """)
        c.execute("CREATE INDEX IF NOT EXISTS idx_push_user ON push_subscriptions(user_hash)")
        c.execute("""
            CREATE TABLE IF NOT EXISTS push_delivery_log (
                id SERIAL PRIMARY KEY,
                endpoint TEXT NOT NULL,
                plan_id INTEGER NOT NULL,
                local_date TEXT NOT NULL,
                local_time TEXT NOT NULL,
                sent_at TEXT NOT NULL,
                UNIQUE(endpoint, plan_id, local_date, local_time)
            )
        """)
        c.execute("CREATE INDEX IF NOT EXISTS idx_push_delivery_plan ON push_delivery_log(plan_id)")
    else:
        c.execute("""
            CREATE TABLE IF NOT EXISTS push_subscriptions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_hash TEXT NOT NULL,
                endpoint TEXT UNIQUE NOT NULL,
                p256dh TEXT NOT NULL,
                auth TEXT NOT NULL,
                timezone TEXT DEFAULT 'Asia/Riyadh',
                lang TEXT DEFAULT 'ar',
                active INTEGER DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        """)
        c.execute("CREATE INDEX IF NOT EXISTS idx_push_user ON push_subscriptions(user_hash)")
        c.execute("""
            CREATE TABLE IF NOT EXISTS push_delivery_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                endpoint TEXT NOT NULL,
                plan_id INTEGER NOT NULL,
                local_date TEXT NOT NULL,
                local_time TEXT NOT NULL,
                sent_at TEXT NOT NULL,
                UNIQUE(endpoint, plan_id, local_date, local_time)
            )
        """)
        c.execute("CREATE INDEX IF NOT EXISTS idx_push_delivery_plan ON push_delivery_log(plan_id)")


def _create_ss_tables(c):
    """Create SymptoSense Smart Account tables if they don't exist."""
    if USE_POSTGRES:
        c.execute("""
            CREATE TABLE IF NOT EXISTS ss_users (
                id SERIAL PRIMARY KEY,
                email TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                password_hash TEXT NOT NULL,
                created_at TEXT NOT NULL,
                last_login TEXT
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS ss_health_profiles (
                user_id INTEGER PRIMARY KEY REFERENCES ss_users(id) ON DELETE CASCADE,
                display_name TEXT DEFAULT '',
                dob TEXT DEFAULT '',
                gender TEXT DEFAULT '',
                height TEXT DEFAULT '',
                weight TEXT DEFAULT '',
                activity_level TEXT DEFAULT '',
                medications TEXT DEFAULT '',
                allergies TEXT DEFAULT '',
                health_conditions TEXT DEFAULT '',
                extra_info TEXT DEFAULT '',
                lang TEXT DEFAULT 'ar',
                updated_at TEXT NOT NULL
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS ss_privacy (
                user_id INTEGER PRIMARY KEY REFERENCES ss_users(id) ON DELETE CASCADE,
                use_in_assistant INTEGER DEFAULT 1,
                use_in_analysis INTEGER DEFAULT 1,
                use_in_calculators INTEGER DEFAULT 1,
                save_chat_history INTEGER DEFAULT 1,
                updated_at TEXT NOT NULL
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS ss_chat_history (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES ss_users(id) ON DELETE CASCADE,
                role TEXT NOT NULL,
                content TEXT NOT NULL,
                timestamp TEXT NOT NULL
            )
        """)
        c.execute("CREATE INDEX IF NOT EXISTS idx_ss_chat_user ON ss_chat_history(user_id)")
    else:
        c.execute("""
            CREATE TABLE IF NOT EXISTS ss_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                password_hash TEXT NOT NULL,
                created_at TEXT NOT NULL,
                last_login TEXT
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS ss_health_profiles (
                user_id INTEGER PRIMARY KEY,
                display_name TEXT DEFAULT '',
                dob TEXT DEFAULT '',
                gender TEXT DEFAULT '',
                height TEXT DEFAULT '',
                weight TEXT DEFAULT '',
                activity_level TEXT DEFAULT '',
                medications TEXT DEFAULT '',
                allergies TEXT DEFAULT '',
                health_conditions TEXT DEFAULT '',
                extra_info TEXT DEFAULT '',
                lang TEXT DEFAULT 'ar',
                updated_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES ss_users(id) ON DELETE CASCADE
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS ss_privacy (
                user_id INTEGER PRIMARY KEY,
                use_in_assistant INTEGER DEFAULT 1,
                use_in_analysis INTEGER DEFAULT 1,
                use_in_calculators INTEGER DEFAULT 1,
                save_chat_history INTEGER DEFAULT 1,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES ss_users(id) ON DELETE CASCADE
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS ss_chat_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                role TEXT NOT NULL,
                content TEXT NOT NULL,
                timestamp TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES ss_users(id) ON DELETE CASCADE
            )
        """)
        c.execute("CREATE INDEX IF NOT EXISTS idx_ss_chat_user ON ss_chat_history(user_id)")


def _migrate_ss_columns(conn, c):
    """Add any new columns to existing ss_ tables."""
    def columns(table):
        if USE_POSTGRES:
            c.execute(
                "SELECT column_name FROM information_schema.columns WHERE table_name = %s",
                (table,),
            )
            return {row[0] for row in c.fetchall()}
        c.execute(f"PRAGMA table_info({table})")
        return {row[1] for row in c.fetchall()}
    try:
        hp_cols = columns("ss_health_profiles")
        for col, default in [("extra_info", ""), ("activity_level", "")]:
            if col not in hp_cols:
                c.execute(f"ALTER TABLE ss_health_profiles ADD COLUMN {col} TEXT DEFAULT '{default}'")
    except Exception:
        pass


def _migrate_feedback(conn, c):
    """Add the comment column to feedback tables created before this feature."""
    if USE_POSTGRES:
        c.execute(
            "SELECT column_name FROM information_schema.columns WHERE table_name = %s",
            ("feedback",),
        )
        existing = {row[0] for row in c.fetchall()}
    else:
        c.execute("PRAGMA table_info(feedback)")
        existing = {row[1] for row in c.fetchall()}
    if "comment" not in existing:
        c.execute("ALTER TABLE feedback ADD COLUMN comment TEXT")


def _migrate_records(conn, c):
    """Add any columns introduced after the table was first created."""
    if USE_POSTGRES:
        c.execute(
            "SELECT column_name FROM information_schema.columns WHERE table_name = %s",
            ("records",),
        )
        existing = {row[0] for row in c.fetchall()}
    else:
        c.execute("PRAGMA table_info(records)")
        existing = {row[1] for row in c.fetchall()}
    for col in ("conditions", "medications"):
        if col not in existing:
            c.execute(f"ALTER TABLE records ADD COLUMN {col} TEXT")


def _migrate_members(conn, c):
    """Add the member_id column to records and blood_tests for existing DBs."""
    def columns(table):
        if USE_POSTGRES:
            c.execute(
                "SELECT column_name FROM information_schema.columns WHERE table_name = %s",
                (table,),
            )
            return {row[0] for row in c.fetchall()}
        c.execute(f"PRAGMA table_info({table})")
        return {row[1] for row in c.fetchall()}
    try:
        for table in ("records", "blood_tests"):
            if "member_id" not in columns(table):
                c.execute(f"ALTER TABLE {table} ADD COLUMN member_id INTEGER DEFAULT 0")
    except Exception:
        pass


def fetchall(sql, params=()):
    """Generic read helper so the dashboard doesn't need a raw DB connection."""
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(sql, params)
        return c.fetchall()
    finally:
        conn.close()


def add_subscriber(user_id, lang="ar"):
    """Registers/updates a user for the daily health tip broadcast (opt-out via unsubscribe())."""
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"INSERT INTO subscribers (user_id, lang, subscribed, added_at) VALUES ({PH},{PH},1,{PH}) "
            f"ON CONFLICT(user_id) DO UPDATE SET lang=excluded.lang, subscribed=1",
            (user_id, lang, datetime.now(timezone.utc).isoformat()),
        )
        conn.commit()
    finally:
        conn.close()


def unsubscribe(user_id):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(f"UPDATE subscribers SET subscribed=0 WHERE user_id={PH}", (user_id,))
        conn.commit()
    finally:
        conn.close()


def get_subscribers():
    """Returns [(user_id, lang), ...] for all currently opted-in users."""
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute("SELECT user_id, lang FROM subscribers WHERE subscribed=1")
        return c.fetchall()
    finally:
        conn.close()


def save_push_subscription(user_id, subscription, timezone_name='Asia/Riyadh', lang='ar'):
    endpoint = str(subscription.get('endpoint') or '').strip()
    keys = subscription.get('keys') or {}
    p256dh = str(keys.get('p256dh') or '').strip()
    auth = str(keys.get('auth') or '').strip()
    if not endpoint or not p256dh or not auth:
        raise ValueError('Invalid push subscription')
    now = datetime.now(timezone.utc).isoformat()
    conn = _conn()
    try:
        c = conn.cursor()
        if USE_POSTGRES:
            c.execute(
                f"INSERT INTO push_subscriptions (user_hash, endpoint, p256dh, auth, timezone, lang, active, created_at, updated_at) VALUES ({PH},{PH},{PH},{PH},{PH},{PH},1,{PH},{PH}) "
                f"ON CONFLICT(endpoint) DO UPDATE SET user_hash=EXCLUDED.user_hash, p256dh=EXCLUDED.p256dh, auth=EXCLUDED.auth, timezone=EXCLUDED.timezone, lang=EXCLUDED.lang, active=1, updated_at=EXCLUDED.updated_at",
                (_hash_user(user_id), endpoint, p256dh, auth, timezone_name or 'Asia/Riyadh', lang or 'ar', now, now),
            )
        else:
            c.execute(
                f"INSERT INTO push_subscriptions (user_hash, endpoint, p256dh, auth, timezone, lang, active, created_at, updated_at) VALUES ({PH},{PH},{PH},{PH},{PH},{PH},1,{PH},{PH}) "
                f"ON CONFLICT(endpoint) DO UPDATE SET user_hash=excluded.user_hash, p256dh=excluded.p256dh, auth=excluded.auth, timezone=excluded.timezone, lang=excluded.lang, active=1, updated_at=excluded.updated_at",
                (_hash_user(user_id), endpoint, p256dh, auth, timezone_name or 'Asia/Riyadh', lang or 'ar', now, now),
            )
        conn.commit()
    finally:
        conn.close()


def delete_push_subscription(user_id, endpoint=None):
    conn = _conn()
    try:
        c = conn.cursor()
        if endpoint:
            c.execute(f"UPDATE push_subscriptions SET active=0, updated_at={PH} WHERE user_hash={PH} AND endpoint={PH}", (datetime.now(timezone.utc).isoformat(), _hash_user(user_id), endpoint))
        else:
            c.execute(f"UPDATE push_subscriptions SET active=0, updated_at={PH} WHERE user_hash={PH}", (datetime.now(timezone.utc).isoformat(), _hash_user(user_id)))
        conn.commit()
    finally:
        conn.close()


def list_push_subscriptions(user_id=None, active_only=True):
    conn = _conn()
    try:
        c = conn.cursor()
        where, params = [], []
        if user_id is not None:
            where.append(f"user_hash={PH}")
            params.append(_hash_user(user_id))
        if active_only:
            where.append("active=1")
        sql = "SELECT user_hash, endpoint, p256dh, auth, timezone, lang, active FROM push_subscriptions"
        if where:
            sql += " WHERE " + " AND ".join(where)
        c.execute(sql, tuple(params))
        rows = c.fetchall()
    finally:
        conn.close()
    return [{'user_hash': r[0], 'endpoint': r[1], 'p256dh': r[2], 'auth': r[3], 'timezone': r[4] or 'Asia/Riyadh', 'lang': r[5] or 'ar', 'active': bool(r[6])} for r in rows]


def mark_push_subscription_inactive(endpoint):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(f"UPDATE push_subscriptions SET active=0, updated_at={PH} WHERE endpoint={PH}", (datetime.now(timezone.utc).isoformat(), endpoint))
        conn.commit()
    finally:
        conn.close()


def list_active_med_plans_for_push():
    import json as _json
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute("SELECT id, user_hash, member_id, med_name, dose, times, start_date, days, active FROM med_plans WHERE active=1")
        rows = c.fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        try:
            times = _json.loads(r[5])
        except Exception:
            times = []
        out.append({'id': r[0], 'user_hash': r[1], 'member_id': r[2], 'med_name': r[3], 'dose': r[4] or '', 'times': [str(x) for x in times], 'start_date': r[6] or '', 'days': r[7], 'active': bool(r[8])})
    return out


def release_push_delivery(endpoint, plan_id, local_date, local_time):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(f"DELETE FROM push_delivery_log WHERE endpoint={PH} AND plan_id={PH} AND local_date={PH} AND local_time={PH}", (endpoint, int(plan_id), local_date, local_time))
        conn.commit()
    finally:
        conn.close()


def claim_push_delivery(endpoint, plan_id, local_date, local_time):
    now = datetime.now(timezone.utc).isoformat()
    conn = _conn()
    try:
        c = conn.cursor()
        try:
            c.execute(f"INSERT INTO push_delivery_log (endpoint, plan_id, local_date, local_time, sent_at) VALUES ({PH},{PH},{PH},{PH},{PH})", (endpoint, int(plan_id), local_date, local_time, now))
            conn.commit()
            return True
        except Exception:
            conn.rollback()
            return False
    finally:
        conn.close()


def log_visit(user_id):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"INSERT INTO visits (user_hash, timestamp) VALUES ({PH},{PH})",
            (_hash_user(user_id), datetime.now(timezone.utc).isoformat()),
        )
        conn.commit()
    except Exception as e:
        import logging
        logging.getLogger("SymptoSense").error(f"log_visit failed for user {user_id}: {e!r}")
    finally:
        conn.close()


def get_usage_stats(days=7):
    conn = _conn()
    try:
        c = conn.cursor()
        since = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
        c.execute("SELECT COUNT(*), COUNT(DISTINCT user_hash) FROM visits")
        total_visits, unique_visitors = c.fetchone()
        c.execute("SELECT COUNT(*), COUNT(DISTINCT user_hash) FROM records")
        total_sessions, unique_users_completed = c.fetchone()
        c.execute(f"SELECT COUNT(*) FROM visits WHERE timestamp >= {PH}", (since,))
        visits_this_period = c.fetchone()[0]
        c.execute(f"SELECT COUNT(*) FROM records WHERE timestamp >= {PH}", (since,))
        sessions_this_period = c.fetchone()[0]
        return {
            "total_visits": total_visits or 0,
            "unique_visitors": unique_visitors or 0,
            "total_sessions": total_sessions or 0,
            "unique_users_completed": unique_users_completed or 0,
            "visits_this_period": visits_this_period or 0,
            "sessions_this_period": sessions_this_period or 0,
            "period_days": days,
        }
    finally:
        conn.close()


def _all_records():
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            "SELECT id, timestamp, lang, age, gender, symptoms, conditions, medications, "
            "duration, severity, urgency FROM records ORDER BY id"
        )
        return c.fetchall()
    finally:
        conn.close()


def export_all_records_csv():
    import csv
    rows = _all_records()
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["id","timestamp","lang","age","gender","symptoms","conditions","medications","duration","severity","urgency"])
    writer.writerows(rows)
    return buf.getvalue()


def export_all_records_xlsx():
    """Export anonymized records as Excel file — ready for Power BI / Excel analysis."""
    try:
        import openpyxl
    except ImportError:
        # fallback to CSV wrapped in BytesIO if openpyxl not installed
        csv_str = export_all_records_csv()
        return io.BytesIO(csv_str.encode("utf-8-sig"))

    rows = _all_records()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SymptoSense Records"
    headers = ["id","timestamp","lang","age","gender","symptoms","conditions","medications","duration","severity","urgency"]
    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    fu_rows = _all_followups()
    if fu_rows:
        ws2 = wb.create_sheet("Followups")
        ws2.append(["id","record_id","timestamp","outcome"])
        for row in fu_rows:
            ws2.append(list(row))

    fb_rows = _all_feedback()
    if fb_rows:
        ws4 = wb.create_sheet("Feedback")
        ws4.append(["id","record_id","timestamp","rating","comment"])
        for row in fb_rows:
            ws4.append(list(row))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _hash_user(user_id) -> str:
    salt = os.environ.get("HASH_SALT", "symptosense")
    return hashlib.sha256(f"{salt}:{user_id}".encode()).hexdigest()[:16]


def get_last_record(user_id):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"SELECT id, timestamp, symptoms, severity, duration, urgency "
            f"FROM records WHERE user_hash={PH} ORDER BY id DESC LIMIT 1",
            (_hash_user(user_id),),
        )
        row = c.fetchone()
    finally:
        conn.close()
    if not row:
        return None
    rec_id, ts, symptoms, severity, duration, urgency = row
    return {
        "id": rec_id,
        "timestamp": ts,
        "symptoms": [s for s in symptoms.split(",") if s],
        "severity": severity,
        "duration": duration,
        "urgency": urgency,
        "days_ago": (datetime.now(timezone.utc) - datetime.fromisoformat(ts)).days,
    }


def get_records(user_id, limit=20, member_id=None):
    """Returns the user's diagnosis records, newest first. Optionally filtered by member_id."""
    conn = _conn()
    try:
        c = conn.cursor()
        if member_id is None:
            c.execute(
                f"SELECT id, timestamp, lang, age, gender, symptoms, duration, severity, urgency, conditions, medications "
                f"FROM records WHERE user_hash={PH} ORDER BY id DESC LIMIT {int(limit)}",
                (_hash_user(user_id),),
            )
        else:
            c.execute(
                f"SELECT id, timestamp, lang, age, gender, symptoms, duration, severity, urgency, conditions, medications "
                f"FROM records WHERE user_hash={PH} AND member_id={PH} ORDER BY id DESC LIMIT {int(limit)}",
                (_hash_user(user_id), int(member_id)),
            )
        rows = c.fetchall()
    finally:
        conn.close()
    return [
        {
            "id": r[0],
            "timestamp": r[1],
            "lang": r[2],
            "age": r[3],
            "gender": r[4],
            "symptoms": [s for s in (r[5] or "").split(",") if s],
            "duration": r[6],
            "severity": r[7],
            "urgency": r[8],
            "conditions": r[9] or "",
            "medications": r[10] or "",
        }
        for r in rows
    ]


def save_profile(user_id, lang, age, gender, conditions, medications, allergies):
    conn = _conn()
    try:
        c = conn.cursor()
        now = datetime.now(timezone.utc).isoformat()
        c.execute(
            f"INSERT INTO profiles (user_hash, lang, age, gender, conditions, medications, allergies, updated_at) "
            f"VALUES ({PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH}) "
            f"ON CONFLICT(user_hash) DO UPDATE SET "
            f"lang=excluded.lang, age=excluded.age, gender=excluded.gender, "
            f"conditions=excluded.conditions, medications=excluded.medications, "
            f"allergies=excluded.allergies, updated_at=excluded.updated_at",
            (_hash_user(user_id), lang, age or "", gender or "", conditions or "", medications or "", allergies or "", now),
        )
        conn.commit()
    finally:
        conn.close()


def load_profile(user_id):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            "SELECT lang, age, gender, conditions, medications, allergies, updated_at "
            "FROM profiles WHERE user_hash=%s" % PH,
            (_hash_user(user_id),),
        )
        row = c.fetchone()
    finally:
        conn.close()
    if not row:
        return None
    return {
        "lang": row[0],
        "age": row[1] or "",
        "gender": row[2] or "",
        "conditions": row[3] or "",
        "medications": row[4] or "",
        "allergies": row[5] or "",
        "updated_at": row[6],
    }


def save_result(user_id, record_id, data):
    """Stores the full analysis result JSON, keyed by (user_hash, record_id)."""
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"INSERT INTO results (user_hash, record_id, data) VALUES ({PH},{PH},{PH}) "
            f"ON CONFLICT(user_hash, record_id) DO UPDATE SET data=excluded.data",
            (_hash_user(user_id), record_id, json.dumps(data, ensure_ascii=False)),
        )
        conn.commit()
    finally:
        conn.close()


def load_result(user_id, record_id):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            "SELECT data FROM results WHERE user_hash=%s AND record_id=%s" % (PH, PH),
            (_hash_user(user_id), record_id),
        )
        row = c.fetchone()
    finally:
        conn.close()
    if not row:
        return None
    try:
        return json.loads(row[0])
    except Exception:
        return None


def save_record(user_id, lang, age, gender, symptoms, duration, severity, urgency, conditions=None, medications=None, member_id=0):
    conn = _conn()
    try:
        c = conn.cursor()
        if USE_POSTGRES:
            c.execute(
                f"INSERT INTO records (user_hash, timestamp, lang, age, gender, symptoms, conditions, medications, duration, severity, urgency, member_id) "
                f"VALUES ({PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH}) RETURNING id",
                (
                    _hash_user(user_id),
                    datetime.now(timezone.utc).isoformat(),
                    lang, age, gender,
                    ",".join(symptoms or []),
                    conditions or "", medications or "",
                    duration, severity, urgency, int(member_id or 0),
                ),
            )
            rec_id = c.fetchone()[0]
        else:
            c.execute(
                f"INSERT INTO records (user_hash, timestamp, lang, age, gender, symptoms, conditions, medications, duration, severity, urgency, member_id) "
                f"VALUES ({PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH})",
                (
                    _hash_user(user_id),
                    datetime.now(timezone.utc).isoformat(),
                    lang, age, gender,
                    ",".join(symptoms or []),
                    conditions or "", medications or "",
                    duration, severity, urgency, int(member_id or 0),
                ),
            )
            rec_id = c.lastrowid
        conn.commit()
        return rec_id
    finally:
        conn.close()


def save_blood_test(user_id, data, member_id=0):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"INSERT INTO blood_tests (user_hash, timestamp, data, member_id) VALUES ({PH},{PH},{PH},{PH})",
            (_hash_user(user_id), datetime.now(timezone.utc).isoformat(),
             json.dumps(data, ensure_ascii=False), int(member_id or 0)),
        )
        conn.commit()
        if USE_POSTGRES:
            c.execute("SELECT lastval()")
            new_id = c.fetchone()[0]
        else:
            new_id = c.lastrowid
        return new_id
    finally:
        conn.close()


def get_blood_tests(user_id, limit=1, member_id=None):
    conn = _conn()
    try:
        c = conn.cursor()
        if member_id is None:
            c.execute(
                f"SELECT id, data, timestamp FROM blood_tests WHERE user_hash={PH} ORDER BY timestamp DESC LIMIT {int(limit)}",
                (_hash_user(user_id),),
            )
        else:
            c.execute(
                f"SELECT id, data, timestamp FROM blood_tests WHERE user_hash={PH} AND member_id={PH} ORDER BY timestamp DESC LIMIT {int(limit)}",
                (_hash_user(user_id), int(member_id)),
            )
        rows = c.fetchall()
    finally:
        conn.close()
    out = []
    for row in rows:
        try:
            out.append({"id": row[0], "data": json.loads(row[1]), "timestamp": row[2] or ""})
        except Exception:
            continue
    return out


def get_blood_test(user_id, blood_id):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"SELECT id, data FROM blood_tests WHERE user_hash={PH} AND id={PH}",
            (_hash_user(user_id), blood_id),
        )
        row = c.fetchone()
    finally:
        conn.close()
    if not row:
        return None
    try:
        return {"id": row[0], "data": json.loads(row[1])}
    except Exception:
        return None


def save_followup(user_id, record_id, outcome):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"INSERT INTO followups (user_hash, record_id, timestamp, outcome) VALUES ({PH},{PH},{PH},{PH})",
            (_hash_user(user_id), record_id, datetime.now(timezone.utc).isoformat(), outcome),
        )
        conn.commit()
    finally:
        conn.close()


def _all_followups():
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            "SELECT id, record_id, timestamp, outcome FROM followups ORDER BY id"
        )
        return c.fetchall()
    finally:
        conn.close()


def save_daily_checkin(user_id, severity):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"INSERT INTO daily_checkins (user_hash, timestamp, severity) VALUES ({PH},{PH},{PH})",
            (_hash_user(user_id), datetime.now(timezone.utc).isoformat(), int(severity)),
        )
        conn.commit()
    finally:
        conn.close()


def get_daily_checkins(user_id, days=7):
    """Returns [(date_str, avg_severity), ...] for the last N days, newest last."""
    since = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"SELECT timestamp, severity FROM daily_checkins "
            f"WHERE user_hash={PH} AND timestamp >= {PH} ORDER BY timestamp",
            (_hash_user(user_id), since),
        )
        rows = c.fetchall()
    finally:
        conn.close()
    by_day = {}
    for ts, sev in rows:
        day = ts[:10]
        by_day.setdefault(day, []).append(sev)
    out = []
    for day in sorted(by_day):
        vals = by_day[day]
        out.append((day, round(sum(vals) / len(vals), 2)))
    return out


def add_med_reminder(user_id, med_name, time_utc, lang="ar"):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"INSERT INTO med_reminders (user_id, med_name, time_utc, lang, active, created) "
            f"VALUES ({PH},{PH},{PH},{PH},1,{PH})",
            (user_id, med_name, time_utc, lang, datetime.now(timezone.utc).isoformat()),
        )
        conn.commit()
    finally:
        conn.close()


def remove_med_reminders(user_id):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(f"UPDATE med_reminders SET active=0 WHERE user_id={PH}", (user_id,))
        conn.commit()
    finally:
        conn.close()


def get_active_med_reminders():
    """Returns [(user_id, med_name, time_utc, lang), ...] for active reminders."""
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            "SELECT user_id, med_name, time_utc, lang FROM med_reminders WHERE active=1 ORDER BY time_utc"
        )
        return c.fetchall()
    finally:
        conn.close()


def save_feedback(user_id, record_id, rating, comment=None):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"INSERT INTO feedback (user_hash, record_id, rating, comment, timestamp) VALUES ({PH},{PH},{PH},{PH},{PH})",
            (_hash_user(user_id), record_id, rating, comment, datetime.now(timezone.utc).isoformat()),
        )
        conn.commit()
    finally:
        conn.close()


def update_feedback_comment(user_id, record_id, comment):
    """Attaches a free-text comment to the latest feedback row for this record."""
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"UPDATE feedback SET comment={PH} WHERE user_hash={PH} AND record_id={PH} "
            f"AND (comment IS NULL OR comment = '')",
            (comment, _hash_user(user_id), record_id),
        )
        conn.commit()
    finally:
        conn.close()


def save_assistant_feedback(user_id, message, rating, reason=None):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"INSERT INTO assistant_feedback (user_hash, message, rating, reason, timestamp) "
            f"VALUES ({PH},{PH},{PH},{PH},{PH})",
            (_hash_user(user_id), message or None, rating, reason,
             datetime.now(timezone.utc).isoformat()),
        )
        conn.commit()
    finally:
        conn.close()


def assistant_feedback_stats():
    """Aggregate assistant feedback for the dashboard."""
    conn = _conn()
    try:
        c = conn.cursor()

        def _count(sql):
            c.execute(sql)
            row = c.fetchone()
            return row[0] if row else 0

        total = _count("SELECT COUNT(*) FROM assistant_feedback")
        useful = _count("SELECT COUNT(*) FROM assistant_feedback WHERE rating = 1")
        partial = _count("SELECT COUNT(*) FROM assistant_feedback WHERE rating = 2")
        not_useful = _count("SELECT COUNT(*) FROM assistant_feedback WHERE rating = 0")
        c.execute(
            "SELECT reason, COUNT(*) FROM assistant_feedback "
            "WHERE rating = 0 AND reason IS NOT NULL AND reason != '' "
            "GROUP BY reason ORDER BY COUNT(*) DESC"
        )
        reasons = c.fetchall()
    finally:
        conn.close()
    satisfaction = int(round((useful + 0.5 * partial) / total * 100)) if total else 0
    return {
        "total": total,
        "useful": useful,
        "partial": partial,
        "not_useful": not_useful,
        "satisfaction": satisfaction,
        "reasons": [{"reason": r, "count": n} for r, n in reasons],
    }


def _all_feedback():
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute("SELECT id, record_id, timestamp, rating, comment FROM feedback ORDER BY id")
        return c.fetchall()
    finally:
        conn.close()


def get_trends(days=7):
    conn = _conn()
    try:
        c = conn.cursor()
        since = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
        c.execute(f"SELECT symptoms FROM records WHERE timestamp >= {PH}", (since,))
        rows = c.fetchall()
    finally:
        conn.close()
    counter = Counter()
    for (symptoms_str,) in rows:
        for s in symptoms_str.split(","):
            s = s.strip()
            if s:
                counter[s] += 1
    return counter, len(rows)


# ---- Family Health Hub & Medication Companion ----

def save_member(user_id, relation, name, age, gender, conditions="", medications="", allergies="", notes=""):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"INSERT INTO family_members (user_hash, relation, name, age, gender, conditions, medications, allergies, notes, created) "
            f"VALUES ({PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH})",
            (_hash_user(user_id), relation, name, age or "", gender or "", conditions or "",
             medications or "", allergies or "", notes or "",
             datetime.now(timezone.utc).isoformat()),
        )
        conn.commit()
        if USE_POSTGRES:
            c.execute("SELECT lastval()")
            return c.fetchone()[0]
        return c.lastrowid
    finally:
        conn.close()


def list_members(user_id):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"SELECT id, relation, name, age, gender, conditions, medications, allergies, notes, created "
            f"FROM family_members WHERE user_hash={PH} ORDER BY id",
            (_hash_user(user_id),),
        )
        rows = c.fetchall()
    finally:
        conn.close()
    return [
        {
            "id": r[0], "relation": r[1] or "", "name": r[2],
            "age": r[3] or "", "gender": r[4] or "",
            "conditions": r[5] or "", "medications": r[6] or "",
            "allergies": r[7] or "", "notes": r[8] or "", "created": r[9] or "",
        }
        for r in rows
    ]


def get_member(user_id, member_id):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"SELECT id, relation, name, age, gender, conditions, medications, allergies, notes, created "
            f"FROM family_members WHERE user_hash={PH} AND id={PH}",
            (_hash_user(user_id), int(member_id)),
        )
        row = c.fetchone()
    finally:
        conn.close()
    if not row:
        return None
    return {
        "id": row[0], "relation": row[1] or "", "name": row[2],
        "age": row[3] or "", "gender": row[4] or "",
        "conditions": row[5] or "", "medications": row[6] or "",
        "allergies": row[7] or "", "notes": row[8] or "", "created": row[9] or "",
    }


def update_member(user_id, member_id, relation, name, age, gender, conditions="", medications="", allergies="", notes=""):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"UPDATE family_members SET relation={PH}, name={PH}, age={PH}, gender={PH}, "
            f"conditions={PH}, medications={PH}, allergies={PH}, notes={PH} "
            f"WHERE user_hash={PH} AND id={PH}",
            (relation, name, age or "", gender or "", conditions or "", medications or "",
             allergies or "", notes or "", _hash_user(user_id), int(member_id)),
        )
        conn.commit()
    finally:
        conn.close()


def delete_member(user_id, member_id):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"DELETE FROM family_members WHERE user_hash={PH} AND id={PH}",
            (_hash_user(user_id), int(member_id)),
        )
        conn.commit()
    finally:
        conn.close()


def save_med_plan(user_id, member_id, med_name, times, dose="", days=None, start_date=None, frequency="daily"):
    import json as _json
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"INSERT INTO med_plans (user_hash, member_id, med_name, dose, times, frequency, start_date, days, active, created) "
            f"VALUES ({PH},{PH},{PH},{PH},{PH},{PH},{PH},{PH},1,{PH})",
            (_hash_user(user_id), int(member_id or 0), med_name, dose or "",
             _json.dumps(times, ensure_ascii=False), frequency,
             start_date or datetime.now(timezone.utc).date().isoformat(),
             int(days) if days else None,
             datetime.now(timezone.utc).isoformat()),
        )
        conn.commit()
        if USE_POSTGRES:
            c.execute("SELECT lastval()")
            return c.fetchone()[0]
        return c.lastrowid
    finally:
        conn.close()


def list_med_plans(user_id, member_id=None, active_only=False):
    import json as _json
    conn = _conn()
    try:
        c = conn.cursor()
        if member_id is None:
            c.execute(
                f"SELECT id, member_id, med_name, dose, times, frequency, start_date, days, active, created "
                f"FROM med_plans WHERE user_hash={PH} ORDER BY id",
                (_hash_user(user_id),),
            )
        else:
            c.execute(
                f"SELECT id, member_id, med_name, dose, times, frequency, start_date, days, active, created "
                f"FROM med_plans WHERE user_hash={PH} AND member_id={PH} ORDER BY id",
                (_hash_user(user_id), int(member_id)),
            )
        rows = c.fetchall()
    finally:
        conn.close()
    out = []
    for r in rows:
        try:
            times = _json.loads(r[4])
        except Exception:
            times = []
        plan = {
            "id": r[0], "member_id": r[1], "med_name": r[2], "dose": r[3] or "",
            "times": times, "frequency": r[5] or "daily", "start_date": r[6] or "",
            "days": r[7], "active": bool(r[8]), "created": r[9] or "",
        }
        if active_only and not plan["active"]:
            continue
        out.append(plan)
    return out


def delete_med_plan(user_id, plan_id):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"UPDATE med_plans SET active=0 WHERE user_hash={PH} AND id={PH}",
            (_hash_user(user_id), int(plan_id)),
        )
        conn.commit()
    finally:
        conn.close()


def log_med_status(user_id, member_id, plan_id, log_date, log_time, status):
    import json as _json
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"DELETE FROM med_logs WHERE plan_id={PH} AND log_date={PH} AND log_time={PH}",
            (int(plan_id), log_date, log_time),
        )
        c.execute(
            f"INSERT INTO med_logs (user_hash, member_id, plan_id, log_date, log_time, status, created) "
            f"VALUES ({PH},{PH},{PH},{PH},{PH},{PH},{PH})",
            (_hash_user(user_id), int(member_id or 0), int(plan_id), log_date, log_time, status,
             datetime.now(timezone.utc).isoformat()),
        )
        conn.commit()
    finally:
        conn.close()


def get_med_logs(user_id, plan_id=None, log_date=None):
    conn = _conn()
    try:
        c = conn.cursor()
        if plan_id is not None and log_date is not None:
            c.execute(
                f"SELECT plan_id, log_date, log_time, status FROM med_logs "
                f"WHERE user_hash={PH} AND plan_id={PH} AND log_date={PH}",
                (_hash_user(user_id), int(plan_id), log_date),
            )
        elif log_date is not None:
            c.execute(
                f"SELECT plan_id, log_date, log_time, status FROM med_logs "
                f"WHERE user_hash={PH} AND log_date={PH}",
                (_hash_user(user_id), log_date),
            )
        else:
            c.execute(
                f"SELECT plan_id, log_date, log_time, status FROM med_logs "
                f"WHERE user_hash={PH}",
                (_hash_user(user_id),),
            )
        rows = c.fetchall()
    finally:
        conn.close()
    return [{"plan_id": r[0], "log_date": r[1], "log_time": r[2], "status": r[3]} for r in rows]


def med_plans_today(user_id):
    """Active plans joined with today's (UTC) log entries."""
    import json as _json
    plans = list_med_plans(user_id, active_only=True)
    today = datetime.now(timezone.utc).date().isoformat()
    logs = get_med_logs(user_id, log_date=today)
    log_map = {}
    for l in logs:
        log_map.setdefault((l["plan_id"], l["log_time"]), l["status"])
    out = []
    for p in plans:
        # duration check
        if p["days"]:
            try:
                end = datetime.strptime(p["start_date"], "%Y-%m-%d").date() + timedelta(days=int(p["days"]))
                if datetime.now(timezone.utc).date() > end:
                    continue
            except Exception:
                pass
        out.append({
            "id": p["id"], "member_id": p["member_id"], "med_name": p["med_name"],
            "dose": p["dose"], "times": p["times"],
            "status": {t: log_map.get((p["id"], t), "") for t in p["times"]},
        })
    return out


def med_adherence(user_id, member_id=None, days=7):
    """Returns {percent, taken, expected, days} for the last N days."""
    import json as _json
    plans = list_med_plans(user_id, member_id=member_id, active_only=False)
    if not plans:
        return {"percent": None, "taken": 0, "expected": 0, "days": days}
    logs = get_med_logs(user_id)
    log_keys = {(l["plan_id"], l["log_date"], l["log_time"]): l["status"] for l in logs}
    today = datetime.now(timezone.utc).date()
    taken = expected = 0
    for p in plans:
        if not p["active"] and p["days"]:
            # skip finished plans for the "expected" count
            continue
        try:
            start = datetime.strptime(p["start_date"], "%Y-%m-%d").date()
        except Exception:
            start = today
        end = today
        if p["days"]:
            try:
                end = min(end, start + timedelta(days=int(p["days"]) - 1))
            except Exception:
                pass
        day_lo = max(start, today - timedelta(days=days - 1))
        for i in range((end - day_lo).days + 1):
            day = day_lo + timedelta(days=i)
            ds = day.isoformat()
            for t in p["times"]:
                expected += 1
                if log_keys.get((p["id"], ds, t)) == "taken":
                    taken += 1
    percent = round((taken * 100.0) / expected, 1) if expected else 0.0
    return {"percent": percent, "taken": taken, "expected": expected, "days": days}


def member_timeline(user_id, member_id, days=30):
    """Merges records + blood tests + med plan events into a reverse-chronological list."""
    import json as _json
    events = []
    since = (datetime.now(timezone.utc) - timedelta(days=int(days))).isoformat()
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            f"SELECT id, timestamp, symptoms FROM records WHERE user_hash={PH} AND member_id={PH} AND timestamp >= {PH}",
            (_hash_user(user_id), int(member_id), since),
        )
        for rid, ts, syms in c.fetchall():
            events.append({
                "date": (ts or "")[:10], "type": "analysis",
                "title": "تحليل أعراض", "en_title": "Symptom analysis",
                "detail": ", ".join([s for s in (syms or "").split(",") if s])[:80],
            })
        c.execute(
            f"SELECT id, timestamp FROM blood_tests WHERE user_hash={PH} AND member_id={PH} AND timestamp >= {PH}",
            (_hash_user(user_id), int(member_id), since),
        )
        for bid, ts in c.fetchall():
            events.append({
                "date": (ts or "")[:10], "type": "blood",
                "title": "فحص CBC", "en_title": "CBC test",
                "detail": "", "id": bid,
            })
    finally:
        conn.close()
    for p in list_med_plans(user_id, member_id=member_id):
        events.append({
            "date": (p["start_date"] or "")[:10], "type": "med",
            "title": "دواء: " + p["med_name"], "en_title": "Medication: " + p["med_name"],
            "detail": ", ".join(p["times"]),
        })
    events.sort(key=lambda e: e["date"], reverse=True)
    return events


# ---- Persistent user/conversation state (survives bot restarts) ----

def save_user_data(user_id, data):
    conn = _conn()
    try:
        c = conn.cursor()
        blob = json.dumps(data, ensure_ascii=False)
        now = datetime.now(timezone.utc).isoformat()
        c.execute(
            f"INSERT INTO user_data (user_id, data, updated_at) VALUES ({PH},{PH},{PH}) "
            f"ON CONFLICT (user_id) DO UPDATE SET data={PH}, updated_at={PH}",
            (user_id, blob, now, blob, now),
        )
        conn.commit()
    finally:
        conn.close()


def load_user_data(user_id):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(f"SELECT data FROM user_data WHERE user_id={PH}", (user_id,))
        row = c.fetchone()
        if not row or not row[0]:
            return {}
        return json.loads(row[0])
    finally:
        conn.close()


def clear_user_data(user_id):
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(f"DELETE FROM user_data WHERE user_id={PH}", (user_id,))
        conn.commit()
    finally:
        conn.close()


def all_user_data():
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute("SELECT user_id, data FROM user_data")
        out = {}
        for user_id, blob in c.fetchall():
            try:
                out[user_id] = json.loads(blob)
            except Exception:
                continue
        return out
    finally:
        conn.close()


def save_conversation(name, key, state):
    conn = _conn()
    try:
        c = conn.cursor()
        name_key = name if name is not None else "default"
        state_blob = json.dumps(state, ensure_ascii=False) if state is not None else None
        key_blob = json.dumps(key, ensure_ascii=False)
        c.execute(
            f"INSERT INTO conversations (name, key, state) VALUES ({PH},{PH},{PH}) "
            f"ON CONFLICT (name, key) DO UPDATE SET state={PH}",
            (name_key, key_blob, state_blob, state_blob),
        )
        conn.commit()
    finally:
        conn.close()


def all_conversations():
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute("SELECT name, key, state FROM conversations")
        out = {}
        for name, key_blob, state_blob in c.fetchall():
            conv_name = None if name == "default" else name
            key = tuple(json.loads(key_blob))
            state = json.loads(state_blob) if state_blob else None
            out.setdefault(conv_name, {})[key] = state
        return out
    finally:
        conn.close()


# ---- SymptoSense Smart Account System ----

import re as _re

def _hash_password(password):
    """Hash password with SHA-256 + salt."""
    salt = os.environ.get("HASH_SALT", "symptosense")
    return hashlib.sha256(f"{salt}:{password}".encode()).hexdigest()


def create_ss_user(email, name, password):
    """Create a new user account. Returns (user_id, error_message)."""
    email = (email or "").strip().lower()
    name = (name or "").strip()
    password = password or ""
    if not email or not _re.match(r'^[^@]+@[^@]+\.[^@]+$', email):
        return None, "invalid_email"
    if not name or len(name) < 2:
        return None, "invalid_name"
    if len(password) < 6:
        return None, "password_too_short"
    conn = _conn()
    try:
        c = conn.cursor()
        now = datetime.now(timezone.utc).isoformat()
        pw_hash = _hash_password(password)
        if USE_POSTGRES:
            c.execute(
                "INSERT INTO ss_users (email, name, password_hash, created_at) VALUES (%s,%s,%s,%s) RETURNING id",
                (email, name, pw_hash, now),
            )
            user_id = c.fetchone()[0]
        else:
            c.execute(
                "INSERT INTO ss_users (email, name, password_hash, created_at) VALUES (?,?,?,?)",
                (email, name, pw_hash, now),
            )
            user_id = c.lastrowid
        conn.commit()
        return user_id, None
    except Exception:
        return None, "email_exists"
    finally:
        conn.close()


def authenticate_ss_user(email, password):
    """Authenticate user. Returns user_id or None."""
    email = (email or "").strip().lower()
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            "SELECT id, password_hash FROM ss_users WHERE email=%s" % PH,
            (email,),
        )
        row = c.fetchone()
        if not row:
            return None
        user_id, stored_hash = row
        if _hash_password(password) != stored_hash:
            return None
        now = datetime.now(timezone.utc).isoformat()
        c.execute("UPDATE ss_users SET last_login=%s WHERE id=%s" % (PH, PH), (now, user_id))
        conn.commit()
        return user_id
    finally:
        conn.close()


def get_ss_user(user_id):
    """Get user info by ID."""
    if not user_id:
        return None
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute("SELECT id, email, name, created_at, last_login FROM ss_users WHERE id=%s" % PH, (int(user_id),))
        row = c.fetchone()
        if not row:
            return None
        return {"id": row[0], "email": row[1], "name": row[2], "created_at": row[3], "last_login": row[4]}
    finally:
        conn.close()


def get_ss_user_by_email(email):
    """Get user info by email."""
    email = (email or "").strip().lower()
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute("SELECT id, email, name FROM ss_users WHERE email=%s" % PH, (email,))
        row = c.fetchone()
        if not row:
            return None
        return {"id": row[0], "email": row[1], "name": row[2]}
    finally:
        conn.close()


def save_health_profile(user_id, data):
    """Save or update health profile for a user."""
    if not user_id:
        return
    conn = _conn()
    try:
        c = conn.cursor()
        now = datetime.now(timezone.utc).isoformat()
        uid = int(user_id)
        fields = {
            "display_name": data.get("display_name", ""),
            "dob": data.get("dob", ""),
            "gender": data.get("gender", ""),
            "height": data.get("height", ""),
            "weight": data.get("weight", ""),
            "activity_level": data.get("activity_level", ""),
            "medications": data.get("medications", ""),
            "allergies": data.get("allergies", ""),
            "health_conditions": data.get("health_conditions", ""),
            "extra_info": data.get("extra_info", ""),
            "lang": data.get("lang", "ar"),
            "updated_at": now,
        }
        all_cols = ["user_id"] + list(fields.keys())
        all_vals = [uid] + list(fields.values())
        if USE_POSTGRES:
            cols = ", ".join(all_cols)
            phs = ", ".join(["%s"] * len(all_cols))
            updates = ", ".join(f"{k}=excluded.{k}" for k in fields.keys())
            c.execute(
                f"INSERT INTO ss_health_profiles ({cols}) VALUES ({phs}) "
                f"ON CONFLICT(user_id) DO UPDATE SET {updates}",
                tuple(all_vals),
            )
        else:
            cols = ", ".join(all_cols)
            phs = ", ".join(["?"] * len(all_cols))
            c.execute(
                f"INSERT OR REPLACE INTO ss_health_profiles ({cols}) VALUES ({phs})",
                tuple(all_vals),
            )
        conn.commit()
    finally:
        conn.close()


def load_health_profile(user_id):
    """Load health profile for a user."""
    if not user_id:
        return None
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            "SELECT display_name, dob, gender, height, weight, activity_level, "
            "medications, allergies, health_conditions, extra_info, lang, updated_at "
            "FROM ss_health_profiles WHERE user_id=%s" % PH,
            (int(user_id),),
        )
        row = c.fetchone()
        if not row:
            return None
        return {
            "display_name": row[0] or "",
            "dob": row[1] or "",
            "gender": row[2] or "",
            "height": row[3] or "",
            "weight": row[4] or "",
            "activity_level": row[5] or "",
            "medications": row[6] or "",
            "allergies": row[7] or "",
            "health_conditions": row[8] or "",
            "extra_info": row[9] or "",
            "lang": row[10] or "ar",
            "updated_at": row[11] or "",
        }
    finally:
        conn.close()


def delete_health_profile(user_id):
    """Delete health profile for a user."""
    if not user_id:
        return
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute("DELETE FROM ss_health_profiles WHERE user_id=%s" % PH, (int(user_id),))
        conn.commit()
    finally:
        conn.close()


def save_privacy_settings(user_id, data):
    """Save privacy settings for a user."""
    if not user_id:
        return
    conn = _conn()
    try:
        c = conn.cursor()
        now = datetime.now(timezone.utc).isoformat()
        use_assistant = 1 if data.get("use_in_assistant", True) else 0
        use_analysis = 1 if data.get("use_in_analysis", True) else 0
        use_calc = 1 if data.get("use_in_calculators", True) else 0
        save_chat = 1 if data.get("save_chat_history", True) else 0
        if USE_POSTGRES:
            c.execute(
                "INSERT INTO ss_privacy (user_id, use_in_assistant, use_in_analysis, use_in_calculators, save_chat_history, updated_at) "
                "VALUES (%s,%s,%s,%s,%s,%s) ON CONFLICT(user_id) DO UPDATE SET "
                "use_in_assistant=excluded.use_in_assistant, use_in_analysis=excluded.use_in_analysis, "
                "use_in_calculators=excluded.use_in_calculators, save_chat_history=excluded.save_chat_history, "
                "updated_at=excluded.updated_at",
                (int(user_id), use_assistant, use_analysis, use_calc, save_chat, now),
            )
        else:
            c.execute(
                "INSERT OR REPLACE INTO ss_privacy (user_id, use_in_assistant, use_in_analysis, use_in_calculators, save_chat_history, updated_at) "
                "VALUES (?,?,?,?,?,?)",
                (int(user_id), use_assistant, use_analysis, use_calc, save_chat, now),
            )
        conn.commit()
    finally:
        conn.close()


def load_privacy_settings(user_id):
    """Load privacy settings for a user."""
    if not user_id:
        return {"use_in_assistant": True, "use_in_analysis": True, "use_in_calculators": True, "save_chat_history": True}
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            "SELECT use_in_assistant, use_in_analysis, use_in_calculators, save_chat_history "
            "FROM ss_privacy WHERE user_id=%s" % PH,
            (int(user_id),),
        )
        row = c.fetchone()
        if not row:
            return {"use_in_assistant": True, "use_in_analysis": True, "use_in_calculators": True, "save_chat_history": True}
        return {
            "use_in_assistant": bool(row[0]),
            "use_in_analysis": bool(row[1]),
            "use_in_calculators": bool(row[2]),
            "save_chat_history": bool(row[3]),
        }
    finally:
        conn.close()


def save_chat_message(user_id, role, content):
    """Save a chat message for a user."""
    if not user_id:
        return
    conn = _conn()
    try:
        c = conn.cursor()
        now = datetime.now(timezone.utc).isoformat()
        if USE_POSTGRES:
            c.execute(
                "INSERT INTO ss_chat_history (user_id, role, content, timestamp) VALUES (%s,%s,%s,%s)",
                (int(user_id), role, content, now),
            )
        else:
            c.execute(
                "INSERT INTO ss_chat_history (user_id, role, content, timestamp) VALUES (?,?,?,?)",
                (int(user_id), role, content, now),
            )
        conn.commit()
    finally:
        conn.close()


def get_chat_history(user_id, limit=50):
    """Get chat history for a user."""
    if not user_id:
        return []
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute(
            "SELECT role, content, timestamp FROM ss_chat_history WHERE user_id=%s ORDER BY id DESC LIMIT %s" % (PH, PH),
            (int(user_id), int(limit)),
        )
        rows = c.fetchall()
    finally:
        conn.close()
    return [{"role": r[0], "content": r[1], "timestamp": r[2]} for r in reversed(rows)]


def clear_chat_history(user_id):
    """Clear chat history for a user."""
    if not user_id:
        return
    conn = _conn()
    try:
        c = conn.cursor()
        c.execute("DELETE FROM ss_chat_history WHERE user_id=%s" % PH, (int(user_id),))
        conn.commit()
    finally:
        conn.close()


def delete_ss_user(user_id):
    """Delete a user account and all associated data."""
    if not user_id:
        return
    conn = _conn()
    try:
        c = conn.cursor()
        uid = int(user_id)
        c.execute("DELETE FROM ss_chat_history WHERE user_id=%s" % PH, (uid,))
        c.execute("DELETE FROM ss_privacy WHERE user_id=%s" % PH, (uid,))
        c.execute("DELETE FROM ss_health_profiles WHERE user_id=%s" % PH, (uid,))
        c.execute("DELETE FROM ss_users WHERE id=%s" % PH, (uid,))
        conn.commit()
    finally:
        conn.close()
